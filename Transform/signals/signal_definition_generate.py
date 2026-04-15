import argparse
from collections import OrderedDict
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import os  # 新增：导入os模块
import sys

current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
if project_root not in sys.path:
    sys.path.insert(0, project_root)
from Config.paths import PIN_TABLE_XLSX, SIGNAL_OUTPUT_XML, ensure_output_dirs

def parse_table(xlsx_path, sheet_name="Signal Pin Table"):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[2]]
    col_index = {h: i for i, h in enumerate(headers)}
    devices = OrderedDict()

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        device = row[col_index["Device"]]
        signal_name = row[col_index["Signal Name"]]
        direction = row[col_index["Direction"]]
        plc_address = row[col_index["PLC Address"]]
        signal_kind = row[col_index["Signal Kind"]] or "Physical"
        pin_desc = row[col_index["Pin / Sub-signal"]] or ""
        value = row[col_index["Value"]]
        meaning = row[col_index["Meaning"]] or ""
        condition = row[col_index["Condition"]] or ""

        if not device or not signal_name or not direction:
            continue

        if device not in devices:
            devices[device] = OrderedDict()
        if signal_name not in devices[device]:
            devices[device][signal_name] = {
                "name": signal_name,
                "type": direction,
                "address": plc_address or "",
                "is_virtual": str(signal_kind).lower() == "virtual",
                "pins": [],
                "selects": []
            }

        sig = devices[device][signal_name]

        if pin_desc:
            # expected format: "<PinName> @ <Address>"
            if "@" in pin_desc:
                desc, addr = [x.strip() for x in pin_desc.split("@", 1)]
            else:
                desc, addr = pin_desc.strip(), ""
            sig["pins"].append({"desc": desc, "address": addr})

        if value is not None or meaning or condition:
            # skip empty “pin-only” lines
            if str(value).strip() != "" or meaning or condition:
                sig["selects"].append({
                    "value": "" if value is None else str(value),
                    "desc": meaning,
                    "condition": condition
                })
    return devices

def build_xml(devices):
    root = ET.Element("SignalMap")
    for device_name, signals in devices.items():
        dev = ET.SubElement(root, "Device", {"name": device_name})
        for sig in signals.values():
            attrs = {
                "Name": sig["name"],
                "Type": sig["type"],
                "Address": sig["address"],
            }
            if sig["is_virtual"]:
                attrs["IsVirtual"] = "true"
            signal = ET.SubElement(dev, "Signal", attrs)
            for pin in sig["pins"]:
                ET.SubElement(signal, "Pin", {
                    "Address": pin["address"],
                    "Desc": pin["desc"],
                })
            for sel in sig["selects"]:
                sel_attrs = {
                    "Value": sel["value"],
                    "Desc": sel["desc"],
                }
                if sel["condition"]:
                    sel_attrs["Condition"] = sel["condition"]
                ET.SubElement(signal, "Select", sel_attrs)
    return root

def indent(elem, level=0):
    i = "\n" + level * "    "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "    "
        for e in elem:
            indent(e, level + 1)
        if not e.tail or not e.tail.strip():
            e.tail = i
    elif level and (not elem.tail or not elem.tail.strip()):
        elem.tail = i

def main():
    parser = argparse.ArgumentParser(description="Generate Signal_Definition.xml from a user-manual-style signal pin table.")
    parser.add_argument("--input", default=str(PIN_TABLE_XLSX))
    parser.add_argument("--sheet", default="Signal Pin Table")
    parser.add_argument("--output", default=str(SIGNAL_OUTPUT_XML))
    args = parser.parse_args()
    ensure_output_dirs()
    # 新增：打印路径，方便调试
    print(f"尝试读取文件：{args.input}")
    if not os.path.exists(args.input):
        print(f"错误：文件不存在！请检查路径是否正确：{args.input}")
        return  # 终止程序，避免崩溃

    devices = parse_table(args.input, args.sheet)
    root = build_xml(devices)
    indent(root)
    tree = ET.ElementTree(root)
    tree.write(args.output, encoding="utf-8", xml_declaration=True)
    print(f"成功生成文件：{args.output}")

if __name__ == "__main__":
    main()